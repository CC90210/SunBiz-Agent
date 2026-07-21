"""test_owner_dob.py — owner Date of Birth: extraction, normalization, validation.

The DOB row is being ADDED to the "UW Sheet 2.5" template (operator, 2026-07-21).
As of that date it is absent from every workbook in Drive — verified across 40
sheets, all 5 tabs, all columns — so there is no live deal to test against yet.

These tests therefore build a SYNTHETIC workbook whose personal block mirrors the
real template (col A = source tag, col B = label, col C = value, anchored at row
76) with a DOB row spliced in. That proves the whole chain — label lookup →
date normalization → sanity validation → lead_data — works the moment the row
appears on a real sheet, and pins the behaviour against regressions.

Run: /srv/sunbiz/ceo-agent/.venv/bin/python -m pytest tests/test_owner_dob.py -q
"""

from __future__ import annotations

import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import pytest
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from mca_lead_scrubber import _date_iso, _dob_iso, build_lead_data  # noqa: E402
from scrubber.uw_sheet_parser import parse_uw_sheet  # noqa: E402

# The live personal block, verified 2026-07-21. Row 76 is its first row; the
# business/home address sub-rows are read by fixed offset beneath their anchor,
# so their relative order matters and is reproduced exactly.
_PERSONAL_BLOCK = [
    ("Jotform", "DBA", "Testco"),
    ("Jotform, Buisness Clear", "Federal Tax ID", "12-3456789"),
    ("Jotform", "Type a entity", "Limited Liability Company"),
    ("Jotform", "Business Address", "100 Example St"),
    ("Jotform", "City", "Austin"),
    ("Jotform", "Zip", "78701"),
    ("Jotform, Experian", "Home Address", "100 Example St"),
    ("Jotform, Experian", "City", "Austin"),
    ("Jotform, Experian", "State (2 Letters ONLY)", "TX"),
    ("Jotform, Experian", "Zip", "78701"),
    ("Jotform", "Email", None),
    ("Jotform", "Phone", None),
    ("Jotform, Experian", "Owner First Name", "Dana"),
    ("Jotform, Experian", "Owner Last Name", "Rivera"),
    ("Jotform, Experian, Personal Clear", "SSN", "000000000"),
]


def _sheet(dob_label: str | None = "Date of Birth", dob_value=None):
    """A minimal UW Sheet 2.5 workbook, optionally carrying a DOB row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "UW Sheet 2.5"
    # Left block (col A label → col B value) — enough for build_lead_data.
    ws["A3"], ws["B3"] = "Business Legal Name", "TESTCO LLC"
    ws["A9"], ws["B9"] = "State", "Texas"

    rows = list(_PERSONAL_BLOCK)
    if dob_label is not None:
        rows.append(("Jotform, Experian", dob_label, dob_value))
    for i, (tag, label, value) in enumerate(rows):
        r = 76 + i
        ws.cell(r, 1, tag)
        ws.cell(r, 2, label)
        if value is not None:
            ws.cell(r, 3, value)
    return wb


def _lead(dob_label="Date of Birth", dob_value=None):
    parsed = parse_uw_sheet(_sheet(dob_label, dob_value))
    return parsed, build_lead_data(
        parsed, {"score": 0, "tier": "good", "reasons": []}, {"name": "UW Sheet_1_Test"}, {}
    )


# ── normalization: every shape the UW team actually types ────────────────────

@pytest.mark.parametrize(
    "raw,expected",
    [
        (datetime(1985, 6, 15), "1985-06-15"),   # a real Excel date cell
        (date(1985, 6, 15), "1985-06-15"),
        ("1985-06-15", "1985-06-15"),            # ISO
        ("06/15/1985", "1985-06-15"),            # US, zero-padded
        ("6/15/1985", "1985-06-15"),             # US, unpadded
        ("06-15-1985", "1985-06-15"),
        ("06.15.1985", "1985-06-15"),
        ("06/15/85", "1985-06-15"),              # 2-digit year
        ("June 15, 1985", "1985-06-15"),
        ("Jun 15, 1985", "1985-06-15"),
        ("15 June 1985", "1985-06-15"),
    ],
)
def test_dob_normalizes_to_iso(raw, expected):
    assert _dob_iso(raw) == expected


def test_pdf_renders_us_format():
    """oasis application-pdf.usDate() turns the stored ISO into MM/DD/YYYY. Pin
    the contract from this side: what we store must be ISO, or usDate passes it
    through untouched and the lender sees the raw string."""
    iso = _dob_iso("6/15/1985")
    assert iso == "1985-06-15"
    y, m, d = iso.split("-")
    assert f"{m}/{d}/{y}" == "06/15/1985"


# ── validation: malformed / implausible input is DROPPED, never passed raw ────

@pytest.mark.parametrize(
    "raw",
    [
        None, "", "   ",
        "n/a", "N/A", "unknown", "tbd", "-",
        "555-123-4567",          # a phone number in the DOB cell
        "not a date",
        "13/45/1985",            # impossible month/day
        45123,                   # a bare Excel serial number
    ],
)
def test_malformed_dob_is_dropped(raw):
    assert _dob_iso(raw) is None


def test_future_dob_rejected():
    tomorrow = (datetime.now(timezone.utc).date() + timedelta(days=1)).isoformat()
    assert _dob_iso(tomorrow) is None


def test_implausible_age_rejected():
    """The two failure modes seen in this data: a business start date typed into
    the DOB cell (too young), and a century-mangled year (too old)."""
    recent = (datetime.now(timezone.utc).date() - timedelta(days=365 * 5)).isoformat()
    assert _dob_iso(recent) is None, "a 5-year-old is not an MCA signer"
    assert _dob_iso("1850-01-01") is None


def test_boundary_ages_accepted():
    today = datetime.now(timezone.utc).date()
    just_18 = today.replace(year=today.year - 18).isoformat()
    assert _dob_iso(just_18) == just_18


def test_dob_never_falls_back_to_raw_text():
    """The D5-class regression guard: a bad DOB must not reach lead_data as raw
    text, because oasis usDate() passes non-ISO strings straight onto the PDF."""
    _, data = _lead(dob_value="sometime in 85")
    assert "owner_dob" not in data


# ── end-to-end through the sheet parser ──────────────────────────────────────

def test_dob_flows_from_sheet_to_lead_data():
    parsed, data = _lead(dob_value=datetime(1985, 6, 15))
    assert parsed["owner_dob"] == "1985-06-15"
    assert data["owner_dob"] == "1985-06-15"
    # The rest of the block must still parse with the new row present.
    assert data["owner_name"] == "Dana Rivera"
    assert data["business_address"] == "100 Example St, Austin 78701"


@pytest.mark.parametrize(
    "label",
    ["Date of Birth", "DOB", "DOB:", "D.O.B.", "Birth Date", "Birthdate",
     "Owner DOB", "Owner Date of Birth", "Signer DOB", "1st Owner DOB"],
)
def test_dob_label_variants(label):
    """Whatever the operator names the new row, it must be picked up without a
    redeploy — label_key() also handles the trailing colon and casing."""
    _, data = _lead(dob_label=label, dob_value="06/15/1985")
    assert data.get("owner_dob") == "1985-06-15", f"label {label!r} not matched"


def test_absent_dob_row_is_not_an_error():
    """Today's reality: no sheet has the row. Parsing must succeed and simply
    omit owner_dob."""
    parsed, data = _lead(dob_label=None)
    assert parsed["owner_dob"] is None
    assert "owner_dob" not in data
    assert data["owner_name"] == "Dana Rivera"


def test_date_iso_still_tolerant_for_non_dob_callers():
    """_date_iso is shared with the TIB path, which legitimately wants a lenient
    parse; only _dob_iso adds the age validation on top."""
    assert _date_iso("2017-03-15") == "2017-03-15"
    assert _date_iso("garbage") is None
