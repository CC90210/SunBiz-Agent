from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from mca_lead_scrubber import build_lead_data  # noqa: E402
from scrubber.ingest import excluded_by_name  # noqa: E402


def test_build_lead_data_exposes_command_centre_aliases() -> None:
    parsed = {
        "business_legal_name": "FROZEN ROPES, INC",
        "dba": "Fundraising University",
        "entity_type": "Corporation",
        "ein": "00-0000000",
        "tib": "2018-06-20 00:00:00",
        "state": "Missouri",
        "owner_name": "Nick Martin",
        "owner_first": "Nick",
        "owner_last": "Martin",
        "owner_dob": "01/02/1980",
        "owner_citizenship": "US Citizen",
        "credit_score": "720",
        "ssn": "111-22-3333",
        "business_address": "PO Box 64",
        "business_city": "Anderson",
        "business_zip": "64831",
        "home_address": "100 Main St",
        "home_city": "Anderson",
        "home_state": "MO",
        "home_zip": "64831",
        "true_revenue_monthly": 98741.94,
        "position_count": 3,
        "leverage_pct": 26.94,
        "previously_submitted": False,
        "counted_funders": [],
        "positions": [],
    }
    result = {"score": 61, "tier": "review", "reasons": [], "decline_reason": None}
    ref = {"name": "UW Sheet_123_FROZEN ROPES", "id": "sheet123"}

    data = build_lead_data(parsed, result, ref, {"version": "test"})

    assert data["legal_name"] == "FROZEN ROPES, INC"
    assert data["owner_name"] == "Nick Martin"
    assert data["owner_dob"] == "1980-01-02"
    assert data["owner_citizenship"] == "US Citizen"
    assert data["owner_ssn_last4"] == "3333"
    assert data["credit_score"] == 720
    assert data["owner_address_line1"] == "100 Main St"
    assert data["business_address_line1"] == "PO Box 64"
    assert data["business_state"] == "Missouri"
    assert data["business_state_code"] == "MO"
    assert data["timezone"] == "America/Chicago"
    assert data["business_start_date"] == "2018-06-20"
    assert data["time_in_business_months"] >= 96
    # Command-Centre tile aliases (daemon-native keys map to what the UI reads).
    assert data["avg_monthly_revenue"] == 98741.94
    assert data["monthly_revenue"] == 98741.94
    assert data["open_mca_positions"] == 3
    assert data["mca_positions"] == 3


def test_drive_discovery_excludes_non_deal_sheet_names() -> None:
    env = {}

    assert excluded_by_name(env, "Contracts Sent - July")
    assert excluded_by_name(env, "Notification Log")
    assert excluded_by_name(env, "Do Not Process")
    assert excluded_by_name(env, "UW Sheet_123_moladds@gmail.com has signed your document")
    assert excluded_by_name(env, "UW Sheet_123_You invited x@y.com to sign Breeze Advance LLC")
    assert excluded_by_name(env, "UW Sheet_123_Breeze Advance LLC - RAY TEX INC Has Been Completed")
    assert not excluded_by_name(env, "UW Sheet_123_FROZEN ROPES")


def test_parser_prefers_numeric_credit_over_link_text() -> None:
    """Codex audit P2 (2026-07-03): truthy link text in the analysis column must not
    shadow a real numeric score typed in the personal block."""
    import openpyxl

    from scrubber.uw_sheet_parser import parse_uw_sheet

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(20, 1, "Credit")
    ws.cell(20, 2, "Link")            # analysis column: Experian link placeholder
    ws.cell(30, 2, "Credit Score")    # personal block: label col B → value col C
    ws.cell(30, 3, "720")
    assert parse_uw_sheet(wb)["credit_score"] == "720"

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.cell(20, 1, "Credit")
    ws2.cell(20, 2, "680")            # numeric in the analysis column wins
    assert parse_uw_sheet(wb2)["credit_score"] == "680"


def test_enricher_notify_and_gate_env_parsing() -> None:
    from uw_lead_enricher import _live_enabled, _max_notify, _notify_enabled

    assert _max_notify({}) == 5                              # default cap
    assert _max_notify({"UW_ENRICH_MAX_NOTIFY": "0"}) == 0   # 0 = unlimited
    assert _max_notify({"UW_ENRICH_MAX_NOTIFY": "junk"}) == 5
    assert _notify_enabled({})                               # notices default ON
    assert not _notify_enabled({"UW_ENRICH_NOTIFY_EZRA": "0"})
    assert not _live_enabled({})                             # loop gated by default
    assert _live_enabled({"UW_ENRICH_READY": "1"})
