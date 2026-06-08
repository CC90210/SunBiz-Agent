"""import_mca_leads.py — bulk-import Adon's MCA web-form lead list into the
SunBiz tenant.

Phase 1 deliverable from Adon's 2026-06-08 architecture brief. The
existing `/api/import/leads` path on the dashboard (leads-import-service.ts)
handles the standard SunBiz lead shape but doesn't model the MCA-specific
fields Adon's spreadsheet carries: multi-position funding history with
named current funders + payment schedules, multi-phone with type/network
metadata, SSN-last-4, EIN, DOB, address.

This importer reads either the xlsx (preferred) or a CSV pre-converted
by the operator, maps the MCA columns into a richer jsonb shape, then
inserts into tenant_records(entity_type='lead') with idempotent dedup
on email|phone|business.

USAGE
-----
  python scripts/import_mca_leads.py \\
      --source-path /path/to/MCA_Webforms_June1-5_clean.xlsx \\
      --tenant-id aa04fa1f-ad6a-44b0-ac4b-2ff5d1067110 \\
      --source-tag adon_handoff_2026-06-08 \\
      --date-range 2026-06-01..2026-06-05

DRY RUN
-------
  python scripts/import_mca_leads.py --source-path X.xlsx --dry-run --limit 10
  # parses + maps + dedup-checks; prints what WOULD be inserted; no writes

SECURITY
--------
- SSN: only LAST 4 digits stored (`data.ssn_last4`). Full SSN is hashed
  and discarded at parse time. Operator fetches full SSN from Adon's
  Salesforce when underwriting needs it.
- DOB: stored as-is (`data.dob`). Lower-tier PII than SSN; needed for
  underwriting (lender age verification).
- EIN: stored as-is (`data.ein`). Business identifier, not personal PII.

DEDUP
-----
Three keys checked in order: email -> phone -> (company + state). First
match wins. Existing leads are SKIPPED — never overwritten. To merge
into an existing lead, use the per-lead update endpoint instead.

OUTPUT
------
Stdout: line-by-line progress (one per row, terse).
state/import_mca_leads.{timestamp}.json: structured report
  {parsed, inserted, skipped_duplicate, skipped_malformed,
   duplicate_keys[], errors[], by_state{}, by_positions{}}
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import hmac
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional


REPO_ROOT = Path(__file__).resolve().parent.parent
STATE_DIR = REPO_ROOT / "state"
LOG_PATH = STATE_DIR / "import_mca_leads.log"

sys.path.insert(0, str(REPO_ROOT / "scripts"))
from _bravo_bootstrap import bootstrap_bravo_path  # noqa: E402

BRAVO_ROOT = bootstrap_bravo_path()

SUNBIZ_TENANT_ID = "aa04fa1f-ad6a-44b0-ac4b-2ff5d1067110"
DEFAULT_STAGE = "hot_lead"  # Adon's intake = hot warm lead; sequence_runner
                            # will move it forward as drips progress.


def _log(msg: str) -> None:
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    line = f"[{datetime.now(timezone.utc).isoformat(timespec='seconds')}] {msg}\n"
    try:
        with LOG_PATH.open("a", encoding="utf-8") as fh:
            fh.write(line)
    except OSError:
        pass
    print(line.rstrip())


# ─────────────────────────────────────────────────────────────────────
# Supabase client
# ─────────────────────────────────────────────────────────────────────


def _supabase():
    try:
        from lib.secret_loader import load_env  # type: ignore
    except Exception:
        return None
    env = load_env()
    url = (env.get("BRAVO_SUPABASE_URL") or "").strip()
    key = (env.get("BRAVO_SUPABASE_SERVICE_ROLE_KEY") or "").strip()
    if not url or not key:
        _log("supabase: missing BRAVO_SUPABASE_URL or SERVICE_ROLE_KEY in .env.agents")
        return None
    try:
        from supabase import create_client
    except ImportError:
        _log("supabase: install via `pip install supabase`")
        return None
    return create_client(url, key)


# ─────────────────────────────────────────────────────────────────────
# Parsing — handles both xlsx (via openpyxl) and CSV
# ─────────────────────────────────────────────────────────────────────


def read_rows(source_path: Path) -> list[dict[str, Any]]:
    """Read a spreadsheet into a list of dicts keyed by header name.
    Handles xlsx via openpyxl, csv natively, and PDF via pdfplumber.
    Headers are normalized to lowercase + stripped + underscored for
    stable mapping."""
    ext = source_path.suffix.lower()
    if ext == ".xlsx":
        return _read_xlsx(source_path)
    if ext == ".csv":
        return _read_csv(source_path)
    if ext == ".pdf":
        return _read_pdf(source_path)
    raise ValueError(f"unsupported file extension: {ext}")


def _read_pdf(path: Path) -> list[dict[str, Any]]:
    """Parse Adon's PDF-export of his MCA spreadsheet. The original xlsx
    is too wide to render on one page, so Excel split the columns across
    multiple page-groups (pages 1-15 = primary contact, 16-30 = phones +
    address, 31-45 = SSN/EIN/birth/revenue, 46-60 = positions + funders).
    Within each page-group, rows are aligned by index — row N on page 1
    corresponds to row N on page 16 etc.

    Strategy: extract tables from every page, group pages by their
    header signature, then zip the row sequences across groups by index.
    Falls back to raw text parsing when pdfplumber's table detection
    can't find a grid.
    """
    try:
        import pdfplumber
    except ImportError as exc:
        raise RuntimeError("pdfplumber not installed — `pip install pdfplumber`") from exc

    # Each page-group has a distinct header signature. We detect which
    # group a page belongs to by inspecting its first row.
    group_signatures = [
        ("primary_contact", {"phone_number", "first_name", "last_name", "company", "email"}),
        ("phones_address",  {"phone2", "address"}),
        ("ids_revenue",     {"city", "state", "zip", "ss", "ein", "revenue"}),
        ("positions_funders", {"positions", "funding_company"}),
    ]

    group_rows: dict[str, list[list[Any]]] = {g[0]: [] for g in group_signatures}
    group_headers: dict[str, list[str]] = {g[0]: [] for g in group_signatures}

    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            for table in tables:
                if not table or len(table) < 2:
                    continue
                headers_raw = table[0]
                headers = [_norm_header(h) for h in headers_raw]
                header_set = {h for h in headers if h}
                # Match against known group signatures (at least 60% header overlap)
                matched_group = None
                best_overlap = 0
                for group_name, sig in group_signatures:
                    overlap = len(sig & header_set)
                    if overlap > best_overlap and overlap >= len(sig) * 0.5:
                        best_overlap = overlap
                        matched_group = group_name
                if not matched_group:
                    continue
                # First seen group headers win — later pages keep the same shape
                if not group_headers[matched_group]:
                    group_headers[matched_group] = headers
                for row in table[1:]:
                    # Skip totally-blank rows
                    if not any((cell or "").strip() for cell in row if cell):
                        continue
                    group_rows[matched_group].append(list(row))

    # Codex audit finding #2 (critical PII safety): row alignment by index
    # is unsafe when pdfplumber misses or duplicates a row in any one
    # group — every subsequent row gets misaligned and SSN/EIN/funder
    # data attaches to the WRONG company silently. Validate row counts
    # across groups; refuse the import unless every group has the same
    # count. Operator can fall back to xlsx/CSV (preferred path).
    populated_counts = {
        name: len(rows) for name, rows in group_rows.items() if rows
    }
    if not populated_counts:
        return []
    counts = list(populated_counts.values())
    if len(set(counts)) > 1:
        raise RuntimeError(
            "PDF parse refused — row counts disagree across page-groups: "
            f"{populated_counts}. Aligning by index would attach the wrong "
            "PII to leads. Convert the source to xlsx or CSV (xlsx -> "
            "Save As -> CSV in Excel) and re-run with --source-path <file>.csv."
        )
    expected_count = counts[0]
    # Also refuse if any expected group is entirely missing (the PDF was
    # truncated or pdfplumber failed on a page-group).
    expected_groups = {name for name, _sig in group_signatures}
    missing = expected_groups - set(populated_counts.keys())
    if missing:
        raise RuntimeError(
            f"PDF parse refused — missing column groups: {sorted(missing)}. "
            "Source PDF may be truncated. Convert to xlsx/CSV and re-run."
        )

    # All groups have identical row counts — alignment is safe to proceed.
    out: list[dict[str, Any]] = []
    for i in range(expected_count):
        merged: dict[str, Any] = {}
        for group_name, _sig in group_signatures:
            headers = group_headers.get(group_name) or []
            rows = group_rows.get(group_name) or []
            row_values = rows[i]
            for col_idx, val in enumerate(row_values):
                if col_idx >= len(headers):
                    continue
                key = headers[col_idx]
                if not key:
                    continue
                final_key = key
                if final_key in merged:
                    suffix = 2
                    while f"{key}_{suffix}" in merged:
                        suffix += 1
                    final_key = f"{key}_{suffix}"
                merged[final_key] = val
        if merged:
            out.append(merged)
    return out


def _read_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl not installed — `pip install openpyxl`") from exc
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_norm_header(h) for h in rows[0]]
    out: list[dict[str, Any]] = []
    for raw in rows[1:]:
        d: dict[str, Any] = {}
        for i, val in enumerate(raw):
            if i < len(headers) and headers[i]:
                # When a header repeats (Phone, Phone2, Phone3 with
                # NumberType+NetworkType triplets), suffix later
                # occurrences so the dict keeps both.
                key = headers[i]
                if key in d:
                    suffix = 2
                    while f"{key}_{suffix}" in d:
                        suffix += 1
                    key = f"{key}_{suffix}"
                d[key] = val
        out.append(d)
    return out


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        rows = list(reader)
    if not rows:
        return []
    headers = [_norm_header(h) for h in rows[0]]
    out: list[dict[str, Any]] = []
    for raw in rows[1:]:
        d: dict[str, Any] = {}
        for i, val in enumerate(raw):
            if i < len(headers) and headers[i]:
                key = headers[i]
                if key in d:
                    suffix = 2
                    while f"{key}_{suffix}" in d:
                        suffix += 1
                    key = f"{key}_{suffix}"
                d[key] = val
        out.append(d)
    return out


def _norm_header(h: Any) -> str:
    if h is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "_", str(h).strip().lower()).strip("_")


# ─────────────────────────────────────────────────────────────────────
# Field mapping — PDF/xlsx columns -> tenant_records.data shape
# ─────────────────────────────────────────────────────────────────────


def parse_money(v: Any) -> Optional[float]:
    if v in (None, ""):
        return None
    if isinstance(v, (int, float)):
        return float(v) if v else None
    s = str(v).strip()
    s = re.sub(r"[\$,]", "", s)
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_int(v: Any) -> Optional[int]:
    if v in (None, ""):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v) if v == int(v) else None
    s = str(v).strip()
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        try:
            return int(float(s))
        except ValueError:
            return None


def normalize_phone(v: Any) -> Optional[str]:
    if v in (None, ""):
        return None
    digits = re.sub(r"\D", "", str(v))
    if len(digits) == 10:
        return f"+1{digits}"
    if len(digits) == 11 and digits.startswith("1"):
        return f"+{digits}"
    return None


def normalize_email(v: Any) -> Optional[str]:
    if v in (None, ""):
        return None
    s = str(v).strip().lower()
    if "@" not in s or " " in s:
        return None
    return s


# US state → IANA timezone. Codex audit finding #8: send_gateway's
# _check_send_window falls back to America/Toronto when lead.data.timezone
# is missing; the importer never set timezone, so a California merchant
# would get an 8am Toronto SMS at 5am PT. Setting timezone at import
# eliminates the gap. States with mixed time zones map to the largest
# population zone (FL/TX/ID/IN/KS/KY/MI/ND/NE/OR/SD/TN — covered below).
STATE_TIMEZONE: dict[str, str] = {
    "AL": "America/Chicago",     "AK": "America/Anchorage",
    "AZ": "America/Phoenix",     "AR": "America/Chicago",
    "CA": "America/Los_Angeles", "CO": "America/Denver",
    "CT": "America/New_York",    "DC": "America/New_York",
    "DE": "America/New_York",    "FL": "America/New_York",
    "GA": "America/New_York",    "HI": "Pacific/Honolulu",
    "IA": "America/Chicago",     "ID": "America/Boise",
    "IL": "America/Chicago",     "IN": "America/Indiana/Indianapolis",
    "KS": "America/Chicago",     "KY": "America/New_York",
    "LA": "America/Chicago",     "MA": "America/New_York",
    "MD": "America/New_York",    "ME": "America/New_York",
    "MI": "America/Detroit",     "MN": "America/Chicago",
    "MO": "America/Chicago",     "MS": "America/Chicago",
    "MT": "America/Denver",      "NC": "America/New_York",
    "ND": "America/Chicago",     "NE": "America/Chicago",
    "NH": "America/New_York",    "NJ": "America/New_York",
    "NM": "America/Denver",      "NV": "America/Los_Angeles",
    "NY": "America/New_York",    "OH": "America/New_York",
    "OK": "America/Chicago",     "OR": "America/Los_Angeles",
    "PA": "America/New_York",    "RI": "America/New_York",
    "SC": "America/New_York",    "SD": "America/Chicago",
    "TN": "America/Chicago",     "TX": "America/Chicago",
    "UT": "America/Denver",      "VA": "America/New_York",
    "VT": "America/New_York",    "WA": "America/Los_Angeles",
    "WI": "America/Chicago",     "WV": "America/New_York",
    "WY": "America/Denver",
}


def derive_timezone(state: Any) -> Optional[str]:
    """Map a US state code to IANA timezone for TCPA-compliant SMS
    windows. Returns None for non-US / unknown so the send-window gate
    can fail-closed for SMS instead of assuming America/Toronto."""
    if state in (None, ""):
        return None
    s = str(state).strip().upper()[:2]
    return STATE_TIMEZONE.get(s)


def parse_positions(v: Any) -> Optional[int]:
    """Map Adon's 'Positions' column ('1st', '2nd', '4th or more') into
    an integer 1-4."""
    if v in (None, ""):
        return None
    s = str(v).strip().lower()
    if "1st" in s or s == "1":
        return 1
    if "2nd" in s or s == "2":
        return 2
    if "3rd" in s or s == "3":
        return 3
    if "4th" in s or "more" in s or s == "4":
        return 4
    n = parse_int(s)
    return n if n is not None else None


def parse_current_funders(v: Any) -> tuple[list[dict[str, Any]], str]:
    """Adon's 'Funding Company' column is a free-text dump like:
       'Bizfund weekly $1,917.00, Forward Financing weekly $1,762.50'
    Parse into a structured array. Returns (parsed_array, original_text).
    Parser is forgiving: any segment it can't parse is kept as-is in the
    `notes` field of the row."""
    if v in (None, ""):
        return [], ""
    raw = str(v).strip()
    if not raw:
        return [], ""

    out: list[dict[str, Any]] = []
    # Split on commas but be careful — funder names sometimes have
    # commas in their dollar amounts. Match the pattern
    # "<name> <frequency> $<amount>" greedily up to the next funder boundary.
    pattern = re.compile(
        r"([A-Za-z0-9&._\-\s/]+?)\s+"
        r"(daily|weekly|biweekly|monthly|\*)\s+"
        r"\$?([\d,]+(?:\.\d+)?)",
        re.IGNORECASE,
    )
    for m in pattern.finditer(raw):
        funder = m.group(1).strip().rstrip(",").rstrip()
        frequency = m.group(2).lower()
        amount_str = m.group(3).replace(",", "")
        try:
            amount = float(amount_str)
        except ValueError:
            continue
        if frequency == "*":
            frequency = "monthly_fee"
        out.append({
            "funder": funder,
            "frequency": frequency,
            "payment": amount,
        })
    return out, raw


def _ssn_pepper() -> Optional[str]:
    """Read the SSN HMAC pepper from .env.agents. When missing, we
    refuse to persist any SSN-derived hash (last4 alone is sufficient
    for display; dedup against historical leads must use email/phone/
    business+state instead until the pepper is configured)."""
    try:
        from lib.secret_loader import load_env  # type: ignore
        env = load_env()
    except Exception:
        return None
    p = (env.get("SSN_HMAC_PEPPER") or "").strip()
    # Refuse short / empty peppers — without ≥32 bytes the HMAC is
    # only marginally better than plain SHA-256 against a brute-force
    # dictionary attack of the 10^9 SSN space.
    return p if len(p) >= 32 else None


def hash_ssn(ssn: Any) -> tuple[Optional[str], Optional[str]]:
    """Return (last4, hmac_hash). Codex audit finding #7: the previous
    implementation stored raw SHA-256(ssn). SSNs have a 10^9 search
    space — trivially reversible offline with a rainbow table or
    GPU-accelerated brute force, so the SHA-256 hash was equivalent to
    storing the SSN in cleartext.

    New behavior:
      - last4 is always returned (safe for display)
      - hash is HMAC-SHA256(pepper, ssn) ONLY when SSN_HMAC_PEPPER is
        configured (≥32 bytes) in .env.agents. Without a pepper, hash
        is None — dedup falls back to email/phone/business+state.

    Set the pepper once with `openssl rand -hex 32 >> .env.agents`
    (with key SSN_HMAC_PEPPER=...) and never rotate without re-importing.
    """
    if ssn in (None, ""):
        return None, None
    digits = re.sub(r"\D", "", str(ssn))
    if len(digits) != 9:
        return None, None
    last4 = digits[-4:]
    pepper = _ssn_pepper()
    if not pepper:
        # No pepper configured — return last4 only; never persist
        # an SSN-derived hash without proper salt.
        return last4, None
    h = hmac.new(
        pepper.encode("utf-8"),
        digits.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    return last4, h


def map_row_to_lead_data(
    row: dict[str, Any],
    source_tag: str,
    source_date_range: str,
) -> Optional[dict[str, Any]]:
    """Map one parsed row into the tenant_records.data jsonb shape.
    Returns None if the row is too sparse to be a usable lead (no name +
    no company + no contact channel)."""
    first_name = (row.get("first_name") or "").strip() or None
    last_name = (row.get("last_name") or "").strip() or None
    name_parts = [p for p in (first_name, last_name) if p]
    name = " ".join(name_parts).strip() or None
    company = (row.get("company") or "").strip() or None
    email = normalize_email(row.get("email"))
    phone_primary = normalize_phone(row.get("phone_number") or row.get("phone"))

    # Reject rows with no name, no company, AND no contact channel.
    if not name and not company and not email and not phone_primary:
        return None

    # Multi-phone — Adon's sheet has Phone2 + NumberType + NetworkType,
    # Phone3 + NumberType + NetworkType. Collect any non-empty phone
    # rows into a `phones` array with type/network metadata.
    phones: list[dict[str, Any]] = []
    for prefix in ("phone2", "phone3"):
        num = normalize_phone(row.get(prefix))
        if num:
            ptype = (row.get(f"numbertype_2" if prefix == "phone2" else "numbertype_3") or "").strip() or None
            pnet = (row.get(f"networktype_2" if prefix == "phone2" else "networktype_3") or "").strip() or None
            phones.append({"number": num, "type": ptype, "network": pnet})

    # Address
    address = (row.get("address") or "").strip() or None
    city = (row.get("city") or "").strip() or None
    state = (row.get("state") or "").strip() or None
    zip_code = str(row.get("zip") or "").strip() or None

    # SSN — last 4 only + hash
    ssn_last4, ssn_hash = hash_ssn(row.get("ss") or row.get("ssn"))
    ein = (row.get("ein") or "").strip() or None
    dob = (row.get("birth_date") or row.get("dob") or "").strip() or None
    if dob and isinstance(row.get("birth_date"), datetime):
        dob = row["birth_date"].date().isoformat()

    # Revenue + MCA positions + current funders
    revenue = parse_money(row.get("revenue") or row.get("annual_revenue"))
    positions = parse_positions(row.get("positions"))
    funders, funders_text = parse_current_funders(row.get("funding_company"))

    data: dict[str, Any] = {
        "name": name,
        "first_name": first_name,
        "last_name": last_name,
        "company": company,
        "business_name": company,
        "contact_name": name,
        "email": email,
        "phone": phone_primary,
        "stage": DEFAULT_STAGE,
        "status": "new",
        "score": 0,
        "source": source_tag,
        "source_date_range": source_date_range,
    }
    if phones:
        data["phones"] = phones
    if address:
        data["address"] = address
    if city:
        data["city"] = city
    if state:
        data["state"] = state
        # Codex finding #8: stamp timezone so send_gateway's
        # _check_send_window enforces local-time TCPA windows, not the
        # Toronto fallback. Unknown states get no timezone (the gate
        # will fail-closed for SMS).
        tz = derive_timezone(state)
        if tz:
            data["timezone"] = tz
    if zip_code:
        data["zip"] = zip_code
    if ssn_last4:
        data["ssn_last4"] = ssn_last4
        if ssn_hash:  # only when HMAC pepper is configured — don't store None
            data["ssn_hash"] = ssn_hash
    if ein:
        data["ein"] = ein
    if dob:
        data["dob"] = dob
    if revenue is not None:
        data["annual_revenue"] = revenue
        data["monthly_revenue"] = revenue / 12 if revenue else None
    if positions is not None:
        data["mca_positions"] = positions
    if funders:
        data["current_funders"] = funders
    if funders_text:
        data["current_funders_text"] = funders_text

    return data


# ─────────────────────────────────────────────────────────────────────
# Dedup — read existing leads ONCE, build in-memory keysets
# ─────────────────────────────────────────────────────────────────────


def fetch_existing_keys(sb, tenant_id: str) -> tuple[set[str], set[str], set[str]]:
    """Pull email/phone/business keys from all existing SunBiz leads so
    the importer can dedup without N round-trips. Returns (emails,
    phones, businesses) as lowercased sets."""
    emails: set[str] = set()
    phones: set[str] = set()
    businesses: set[str] = set()
    page_size = 1000
    offset = 0
    while True:
        try:
            r = (
                sb.table("tenant_records")
                .select("data", count="exact")
                .eq("tenant_id", tenant_id)
                .eq("entity_type", "lead")
                .range(offset, offset + page_size - 1)
                .execute()
            )
        except Exception as exc:  # noqa: BLE001
            _log(f"fetch_existing_keys: query failed at offset={offset}: {exc}")
            break
        rows = r.data or []
        for row in rows:
            d = row.get("data") or {}
            e = (d.get("email") or "").strip().lower()
            if e:
                emails.add(e)
            p = (d.get("phone") or "").strip()
            if p:
                phones.add(p)
            c = (d.get("company") or d.get("business_name") or "").strip().lower()
            s = (d.get("state") or "").strip().lower()
            if c:
                businesses.add(f"{c}|{s}")
        if len(rows) < page_size:
            break
        offset += page_size
    return emails, phones, businesses


# ─────────────────────────────────────────────────────────────────────
# Main import path
# ─────────────────────────────────────────────────────────────────────


def run_import(
    source_path: Path,
    tenant_id: str,
    source_tag: str,
    source_date_range: str,
    dry_run: bool = False,
    limit: Optional[int] = None,
) -> dict[str, Any]:
    rows = read_rows(source_path)
    _log(f"parsed {len(rows)} rows from {source_path.name}")
    if limit:
        rows = rows[:limit]
        _log(f"limit applied -> processing first {len(rows)} rows")

    sb = _supabase()
    if not sb and not dry_run:
        return {"ok": False, "error": "supabase_unavailable"}

    if sb:
        existing_emails, existing_phones, existing_businesses = fetch_existing_keys(sb, tenant_id)
        _log(
            f"existing keys loaded: emails={len(existing_emails)} "
            f"phones={len(existing_phones)} businesses={len(existing_businesses)}"
        )
    else:
        existing_emails, existing_phones, existing_businesses = set(), set(), set()

    seen_emails: set[str] = set()
    seen_phones: set[str] = set()
    seen_businesses: set[str] = set()

    to_insert: list[dict[str, Any]] = []
    skipped_dupe = 0
    skipped_malformed = 0
    by_state: dict[str, int] = {}
    by_positions: dict[str, int] = {}
    dupe_keys_sample: list[str] = []
    malformed_sample: list[str] = []

    for row in rows:
        data = map_row_to_lead_data(row, source_tag, source_date_range)
        if not data:
            skipped_malformed += 1
            if len(malformed_sample) < 10:
                malformed_sample.append(
                    f"row#{rows.index(row)}: no name/company/contact"
                )
            continue

        # Dedup
        e = (data.get("email") or "").strip().lower()
        p = (data.get("phone") or "").strip()
        c = (data.get("company") or "").strip().lower()
        s = (data.get("state") or "").strip().lower()
        bkey = f"{c}|{s}" if c else ""

        is_dupe = False
        dupe_key = ""
        if e and (e in existing_emails or e in seen_emails):
            is_dupe = True
            dupe_key = f"email:{e}"
        elif p and (p in existing_phones or p in seen_phones):
            is_dupe = True
            dupe_key = f"phone:{p}"
        elif bkey and (bkey in existing_businesses or bkey in seen_businesses):
            is_dupe = True
            dupe_key = f"business:{bkey}"

        if is_dupe:
            skipped_dupe += 1
            if len(dupe_keys_sample) < 20:
                dupe_keys_sample.append(dupe_key)
            continue

        if e:
            seen_emails.add(e)
        if p:
            seen_phones.add(p)
        if bkey:
            seen_businesses.add(bkey)

        # Telemetry counters
        st = data.get("state") or "(unknown)"
        by_state[st] = by_state.get(st, 0) + 1
        pos = data.get("mca_positions")
        pos_key = str(pos) if pos else "(unknown)"
        by_positions[pos_key] = by_positions.get(pos_key, 0) + 1

        to_insert.append({
            "tenant_id": tenant_id,
            "entity_type": "lead",
            "data": data,
        })

    _log(
        f"map: insertable={len(to_insert)} duplicate={skipped_dupe} "
        f"malformed={skipped_malformed}"
    )

    inserted = 0
    insert_errors: list[str] = []

    if dry_run:
        _log("DRY RUN — no writes performed")
    elif to_insert:
        # Batch in chunks of 200 to keep payloads bounded
        chunk_size = 200
        for i in range(0, len(to_insert), chunk_size):
            chunk = to_insert[i : i + chunk_size]
            try:
                r = sb.table("tenant_records").insert(chunk).execute()
                inserted += len(r.data or [])
            except Exception as exc:  # noqa: BLE001
                msg = f"chunk {i // chunk_size}: {exc}"
                insert_errors.append(msg)
                _log(f"insert error: {msg}")

    report = {
        "ok": True,
        "dry_run": dry_run,
        "parsed": len(rows),
        "insertable": len(to_insert),
        "inserted": inserted,
        "skipped_duplicate": skipped_dupe,
        "skipped_malformed": skipped_malformed,
        "by_state": by_state,
        "by_positions": by_positions,
        "duplicate_keys_sample": dupe_keys_sample,
        "malformed_sample": malformed_sample,
        "insert_errors": insert_errors,
        "source_path": str(source_path),
        "source_tag": source_tag,
        "source_date_range": source_date_range,
        "tenant_id": tenant_id,
        "completed_at": datetime.now(timezone.utc).isoformat(),
    }

    # Persist report
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    report_path = STATE_DIR / f"import_mca_leads.{ts}.json"
    with report_path.open("w", encoding="utf-8") as fh:
        json.dump(report, fh, indent=2)
    _log(f"report written to {report_path}")

    return report


# ─────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────


def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Bulk-import MCA leads into SunBiz tenant")
    parser.add_argument("--source-path", required=True, help="Path to .xlsx or .csv")
    parser.add_argument("--tenant-id", default=SUNBIZ_TENANT_ID)
    parser.add_argument(
        "--source-tag",
        default=f"adon_handoff_{datetime.now().strftime('%Y-%m-%d')}",
        help="data.source value stamped on every imported lead",
    )
    parser.add_argument(
        "--date-range",
        default="",
        help="data.source_date_range stamp (e.g. 2026-06-01..2026-06-05)",
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="parse + dedup-check; no writes")
    parser.add_argument("--limit", type=int, default=None,
                        help="cap rows processed (for smoke testing)")
    parser.add_argument("--json", action="store_true",
                        help="print final report as JSON")
    args = parser.parse_args(argv)

    source = Path(args.source_path).expanduser().resolve()
    if not source.exists():
        print(f"ERROR: source path not found: {source}", file=sys.stderr)
        return 2

    try:
        report = run_import(
            source_path=source,
            tenant_id=args.tenant_id,
            source_tag=args.source_tag,
            source_date_range=args.date_range,
            dry_run=args.dry_run,
            limit=args.limit,
        )
    except Exception as exc:  # noqa: BLE001
        print(f"FATAL: {exc}", file=sys.stderr)
        return 1

    if args.json:
        print(json.dumps(report, indent=2))
    else:
        print(
            f"\nIMPORT REPORT — {source.name}\n"
            f"  parsed:             {report.get('parsed', 0)}\n"
            f"  insertable:         {report.get('insertable', 0)}\n"
            f"  inserted:           {report.get('inserted', 0)}\n"
            f"  skipped_duplicate:  {report.get('skipped_duplicate', 0)}\n"
            f"  skipped_malformed:  {report.get('skipped_malformed', 0)}\n"
            f"  dry_run:            {report.get('dry_run', False)}\n"
        )
        if report.get("insert_errors"):
            print(f"  insert_errors ({len(report['insert_errors'])}):")
            for e in report["insert_errors"][:5]:
                print(f"    - {e}")
    return 0 if report.get("ok") else 1


if __name__ == "__main__":
    sys.exit(main())
